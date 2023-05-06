from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment


webpage = 'https://crypto.com/price'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url=webpage, headers=headers)

page = urlopen(req)

soup = BeautifulSoup(page, 'html.parser')

print(soup.title.text)

crypto_rows = soup.findAll('tr')

for row in crypto_rows:
    td = row.findAll('td')
    if td:
        #print(td[2].text)
        p = row.findAll('p')
        if p:
            print(f"Name: ", p[0].text)
        span = row.findAll('span')
        if span:
            print(f"Symbol: ", span[3].text)
        div = row.findAll('div')
        if div:
            print(f"Current Price: ", div[6].text)
        print(f"Percent Change :", td[4].text)


wb = xl.Workbook()
ws = wb.active
ws.title = 'Cryptocurrency Report'

ws['A1'] = 'Name'
ws['B1'] = "Symbol"
ws['C1'] = "Current Price"
ws['D1'] = "% Change (Last 24 Hours)"
ws['E1'] = "Corresponding Price"

row_index = 2

for row in crypto_rows[1:6]:
    td = row.findAll('td')
    p = row.findAll('p')
    span = row.findAll('span')
    div = row.findAll('div')
    if td:
        if p:
            name = p[0].text
        if span:
            symbol = span[3].text
        if div:
            current_price = int(float(div[6].text.replace(",","").replace("$","")))
            current_price2 = div[6].text
        percent_change = int(float(td[4].text.replace("-","").replace("+","").replace("%","")))
        percent_change2 = td[4].text
        multiplier = percent_change/100
        corresponding_price = round(current_price/(1+multiplier),2)

       
        ws.cell(row=row_index, column=1, value=name)
        ws.cell(row=row_index, column=2, value=symbol)
        ws.cell(row=row_index, column=3, value=current_price2)
        ws.cell(row=row_index, column=4, value=percent_change2)
        ws.cell(row=row_index, column=5, value=corresponding_price)



        row_index += 1
        


header_font = Font(name="Times New Roman",size=14, bold=True, color='ffffff')
for cell in ws[1:1]:
    cell.font = header_font
    cell.fill = PatternFill(fill_type='solid',start_color='1ba13f',end_color='07db40')
text_font = Font(color='e8c413',size=12,name='Lato')
for cell in ws[2:2]:
    cell.font = text_font
for cell in ws[3:3]:
    cell.font = text_font
for cell in ws[4:4]:
    cell.font = text_font
for cell in ws[5:5]:
    cell.font = text_font
for cell in ws[6:6]:
    cell.font = text_font



ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 30

ws.row_dimensions[1].height = 50
wrap_style = Alignment(wrap_text=True)
ws['A1'].alignment = wrap_style
ws['A1'].alignment = Alignment(horizontal='center',vertical='center',)
ws['B1'].alignment = wrap_style
ws['B1'].alignment = Alignment(horizontal='center',vertical='center',)
ws['C1'].alignment = wrap_style
ws['C1'].alignment = Alignment(horizontal='center',vertical='center',)
ws['D1'].alignment = wrap_style
ws['D1'].alignment = Alignment(horizontal='center',vertical='center',)
ws['E1'].alignment = wrap_style
ws['E1'].alignment = Alignment(horizontal='center',vertical='center',)


wb.save('CryptoCurrencyReport.xlsx')



import keys
from twilio.rest import Client
    
client = Client(keys.account_sid, keys.auth_token)

TwilioNumber = "+15075981730"
myphone = '+18328298125'

textmsg_body = f'Buy {name} now! The price has changed!'
difference = int(corresponding_price) - int(current_price)
if (name == 'Bitcoin' or name == 'Ethereum') and (difference > 5 or difference < -5):
        textmsg = client.messages.create(to=myphone, from_=TwilioNumber, body=textmsg_body)
